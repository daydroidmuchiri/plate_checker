import os
import openpyxl
import telegram
from telegram.ext import Updater,  MessageHandler, Filters

# Load the Excel sheet
file_path = "data.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Get a list of all the values in Column A
values = [cell.value for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row]

# Define the bot's behavior when a message is received
def message_handler(update, context):
    message = update.message.text
    if message in values:
        context.bot.send_message(chat_id=update.effective_chat.id, text="registered")
    else:
        context.bot.send_message(chat_id=update.effective_chat.id, text="not registered")

# Set up the bot and start polling for messages
TOKEN = "6081894512:AAG9pdS6tARvHZltY4DXha4d0bW9YzOEX1k"
updater = Updater(TOKEN, use_context=True)
dp = updater.dispatcher
dp.add_handler(MessageHandler(Filters.text, message_handler))
updater.start_polling()
